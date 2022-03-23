# -*- coding:utf-8 -*-
from bs4 import BeautifulSoup
import os
import re
import openpyxl as ws
import sys

def endWith(file,*endstring):
    array = map(file.endswith,endstring)
    if True in array:
        return True
    else:
        return False

def openFile(InputFile,OutFile):
        soup = BeautifulSoup(open(InputFile, mode='r', encoding='utf-8'), 'lxml')
        get_detail(soup,OutFile)
            
def has_border_but_no_class(tag):
    return tag.has_attr('border') and not tag.has_attr('class')

def get_detail(soup,OutFile):
    vulnerabilities_list = []
    #Count all the detect url
    summary_Block = soup.select(".ax-scan-summary")
    tables = soup.find_all(has_border_but_no_class)
    alerts_start = 0
    alerts_Count_loop = 0
    for i in range(len(summary_Block)):
        url = soup.select(".ax-scan-summary > tbody > tr:nth-of-type(3) > td:nth-of-type(2)")[i].string
        alerts_Count = soup.select(".ax-alerts-distribution > tr:nth-of-type(1) > td:nth-of-type(2)")[i].string
        alerts_Count_loop += int(alerts_Count)
        for j in range(alerts_start,alerts_Count_loop):
            scan_url = url
            vl_path = tables[j].select('tr > td > b')[0].string.strip()
            vl_name = tables[j].select('tr:nth-of-type(2) > td > b')[1].string.strip()
            vl_severity = tables[j].select('tr:nth-of-type(3) > td:nth-of-type(2)')[0].string.strip()
            vl_detail = tables[j].select('tr:nth-of-type(7) > td:nth-of-type(2)')[0].get_text().strip()
            vl_post = tables[j].select('tr:nth-of-type(8) > td')[0].string
            vulnerabilities = {}
            vulnerabilities['url'] = scan_url
            vulnerabilities['path'] = vl_path
            vulnerabilities['name'] = vl_name
            vulnerabilities['severity'] = vl_severity
            vulnerabilities['detail'] = vl_detail
            vulnerabilities['post'] = vl_post
            vulnerabilities_list.append(vulnerabilities)
        alerts_start += int(alerts_Count)
    write_xlsx(vulnerabilities_list,OutFile)

def write_xlsx(vulnerabilities_list,OutFile):
    wb = ws.load_workbook(OutFile)
    sheet1 = wb['Sheet']
    num = sheet1.max_row
    for i in range(len(vulnerabilities_list)):
        sheet1.cell(row=num+i+1, column=1, value=vulnerabilities_list[i]['url'])
        sheet1.cell(row=num+i+1, column=2, value=vulnerabilities_list[i]['name'])
        sheet1.cell(row=num+i+1, column=3, value=vulnerabilities_list[i]['path'])
        sheet1.cell(row=num+i+1, column=4, value=vulnerabilities_list[i]['severity'])
        sheet1.cell(row=num+i+1, column=5, value=vulnerabilities_list[i]['post'])
        sheet1.cell(row=num+i+1, column=6, value=vulnerabilities_list[i]['detail'])
    wb.save(OutFile)

def creat_xlsx(OutFile):
    s = 0
    wb = ws.Workbook()
    ws1 = wb.active
    word=['风险目标','风险名称','风险地址','风险等级','风险请求','风险详细']
    for i in word:
        s = s + 1
        ws1.cell(row =1,column = s,value = i)
    wb.save(OutFile)

def main():
    print("Converting")
    creat_xlsx(sys.argv[2])
    openFile(sys.argv[1],sys.argv[2])
    print("Completed！")

if __name__ == '__main__':
    main()
